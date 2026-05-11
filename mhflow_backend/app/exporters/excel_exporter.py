"""MHFlow Excel报告导出器

将热平衡计算结果导出为Excel格式的报告。
"""
import io
import logging
from typing import Dict, Any, Optional
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 库不可用，Excel导出功能将不可用")


class ExcelExporter:
    """
    Excel报告导出器

    将热平衡计算结果生成结构化的Excel工作簿，
    包含概览、系统性能、元件结果、节点参数等多个工作表。
    """

    # 样式常量
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
    BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def __init__(self, results: Dict[str, Any], model_data: Optional[Dict[str, Any]] = None):
        """
        初始化Excel导出器

        参数:
            results: 求解结果
            model_data: 模型数据
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl 库不可用，请安装: pip install openpyxl")

        self.results = results or {}
        self.model_data = model_data or {}

    def export(self, output_path: Optional[str] = None) -> bytes:
        """
        导出Excel报告

        参数:
            output_path: 输出文件路径 (为空则返回字节流)

        返回:
            Excel字节数据
        """
        wb = Workbook()

        # 创建各工作表
        self._create_overview_sheet(wb.active)  # 使用默认sheet作为概览
        self._create_performance_sheet(wb.create_sheet("系统性能"))
        self._create_components_sheet(wb.create_sheet("元件结果"))
        self._create_nodes_sheet(wb.create_sheet("节点参数"))
        self._create_model_data_sheet(wb.create_sheet("模型数据"))

        # 保存到字节流
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_data = buffer.getvalue()
        buffer.close()

        if output_path:
            import os
            os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(excel_data)
            logger.info(f"Excel报告已保存: {output_path}")

        return excel_data

    def _style_header_row(self, ws, row_idx: int, col_count: int):
        """为表头行应用样式"""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGNMENT
            cell.border = self.BORDER

    def _style_data_cell(self, cell, is_number: bool = False):
        """为数据单元格应用样式"""
        cell.border = self.BORDER
        cell.alignment = self.CENTER_ALIGNMENT if is_number else self.LEFT_ALIGNMENT

    def _auto_width(self, ws, min_width: int = 10, max_width: int = 50):
        """自动调整列宽"""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    max_length = max(max_length, val_len)
                except Exception:
                    pass
            adjusted = min(max(min_width, max_length + 2), max_width)
            ws.column_dimensions[col_letter].width = adjusted

    def _create_overview_sheet(self, ws):
        """创建概览工作表"""
        ws.title = "概览"

        # 标题
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = "MHFlow 热平衡计算报告"
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = self.CENTER_ALIGNMENT
        ws.row_dimensions[1].height = 30

        # 基本信息
        info = [
            ("模型名称", self.model_data.get("name", "Unknown Model")),
            ("模型描述", self.model_data.get("description", "")),
            ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("收敛状态", "是" if self.results.get("converged", False) else "否"),
            ("迭代次数", self.results.get("iteration_count", 0)),
            ("元件数量", len(self.results.get("components", {}))),
        ]

        ws.append([])  # 空行
        ws.append(["项目", "值"])
        self._style_header_row(ws, 3, 2)

        for label, value in info:
            ws.append([label, value])
            row = ws.max_row
            self._style_data_cell(ws.cell(row=row, column=1))
            self._style_data_cell(ws.cell(row=row, column=2))
            ws.cell(row=row, column=1).font = Font(bold=True)

        self._auto_width(ws)

    def _create_performance_sheet(self, ws):
        """创建系统性能工作表"""
        performance = self.results.get("system_performance", {})

        ws.append(["系统性能指标"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = self.CENTER_ALIGNMENT
        ws.row_dimensions[1].height = 25

        ws.append([])
        ws.append(["参数", "数值", "单位"])
        self._style_header_row(ws, 3, 3)

        perf_items = [
            ("发电功率", "w_electrical_mw", "MW", 2),
            ("汽轮机内功率", "w_turbine_internal_mw", "MW", 2),
            ("锅炉吸热量", "q_boiler_mw", "MW", 2),
            ("热耗率", "heat_rate_kj_kwh", "kJ/kWh", 2),
            ("锅炉效率", "eta_boiler", "%", 4),
            ("厂用电效率", "eta_plant", "%", 4),
            ("发电煤耗率", "coal_consumption_rate_g_kwh", "g/kWh", 2),
            ("汽耗率", "steam_rate_kg_kwh", "kg/kWh", 4),
            ("主蒸汽流量", "main_steam_flow_t_h", "t/h", 2),
            ("年发电量", "annual_generation_mwh", "MWh", 0),
            ("年耗煤量", "annual_coal_tons", "tons", 0),
        ]

        for label, key, unit, precision in perf_items:
            value = performance.get(key, 0)
            if key in ("eta_boiler", "eta_plant"):
                value = value * 100  # 转换为百分比
            fmt_val = f"{value:.{precision}f}" if isinstance(value, (int, float)) else str(value)
            ws.append([label, fmt_val, unit])
            row = ws.max_row
            for col in range(1, 4):
                self._style_data_cell(ws.cell(row=row, column=col), is_number=(col == 2))

        self._auto_width(ws)

    def _create_components_sheet(self, ws):
        """创建元件结果工作表"""
        components = self.results.get("components", {})

        ws.append(["元件计算结果"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = self.CENTER_ALIGNMENT
        ws.row_dimensions[1].height = 25

        ws.append([])
        ws.append(["元件名称", "元件类型", "参数", "数值"])
        self._style_header_row(ws, 3, 4)

        for comp_name, comp_data in components.items():
            comp_type = comp_data.get("component_type", "")
            results_data = comp_data.get("results", {})

            if not results_data:
                ws.append([comp_name, comp_type, "—", "—"])
                row = ws.max_row
                for col in range(1, 5):
                    self._style_data_cell(ws.cell(row=row, column=col))
                continue

            first_row = True
            for key, value in results_data.items():
                if isinstance(value, float):
                    fmt_val = f"{value:.6g}"
                elif isinstance(value, list):
                    fmt_val = f"[{len(value)} items]"
                else:
                    fmt_val = str(value)

                name_col = comp_name if first_row else ""
                type_col = comp_type if first_row else ""
                ws.append([name_col, type_col, key, fmt_val])
                row = ws.max_row
                for col in range(1, 5):
                    self._style_data_cell(ws.cell(row=row, column=col), is_number=(col == 4))
                first_row = False

        self._auto_width(ws)

    def _create_nodes_sheet(self, ws):
        """创建节点参数工作表"""
        components = self.results.get("components", {})

        ws.append(["节点热力参数"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = self.CENTER_ALIGNMENT
        ws.row_dimensions[1].height = 25

        ws.append([])
        headers = ["节点名称", "所属元件", "端口类型", "压力 P (MPa)", "温度 T (°C)", "比焓 H (kJ/kg)", "比熵 S (kJ/kg·K)", "质量流量 M (kg/s)"]
        ws.append(headers)
        self._style_header_row(ws, 3, len(headers))

        for comp_name, comp_data in components.items():
            # 出口端口
            for port in comp_data.get("outlet_ports", []):
                ws.append([
                    f"{comp_name}.{port.get('name', '')}",
                    comp_name,
                    "出口",
                    f"{port.get('p', 0):.4f}",
                    f"{port.get('t', 0):.2f}",
                    f"{port.get('h', 0):.2f}",
                    f"{port.get('s', 0):.4f}",
                    f"{port.get('m', 0):.4f}",
                ])
                row = ws.max_row
                for col in range(1, len(headers) + 1):
                    self._style_data_cell(ws.cell(row=row, column=col), is_number=(col >= 4))

            # 入口端口
            for port in comp_data.get("inlet_ports", []):
                ws.append([
                    f"{comp_name}.{port.get('name', '')}",
                    comp_name,
                    "入口",
                    f"{port.get('p', 0):.4f}",
                    f"{port.get('t', 0):.2f}",
                    f"{port.get('h', 0):.2f}",
                    f"{port.get('s', 0):.4f}",
                    f"{port.get('m', 0):.4f}",
                ])
                row = ws.max_row
                for col in range(1, len(headers) + 1):
                    self._style_data_cell(ws.cell(row=row, column=col), is_number=(col >= 4))

        self._auto_width(ws)

    def _create_model_data_sheet(self, ws):
        """创建模型数据工作表"""
        ws.append(["模型原始数据"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = self.CENTER_ALIGNMENT
        ws.row_dimensions[1].height = 25

        ws.append([])
        ws.append(["键", "值"])
        self._style_header_row(ws, 3, 2)

        # 简单展平模型数据的前两层
        def flatten(data, prefix=""):
            items = []
            if isinstance(data, dict):
                for k, v in data.items():
                    key = f"{prefix}.{k}" if prefix else k
                    if isinstance(v, (str, int, float, bool)) or v is None:
                        items.append((key, str(v)))
                    elif isinstance(v, dict) and not v:
                        items.append((key, "{}"))
                    elif isinstance(v, list) and not v:
                        items.append((key, "[]"))
                    elif isinstance(v, dict):
                        items.extend(flatten(v, key))
                    elif isinstance(v, list):
                        for i, item in enumerate(v):
                            if isinstance(item, dict):
                                items.extend(flatten(item, f"{key}[{i}]"))
                            else:
                                items.append((f"{key}[{i}]", str(item)))
            return items

        flat_items = flatten(self.model_data)
        # 限制行数避免文件过大
        for key, value in flat_items[:5000]:
            ws.append([key, value])
            row = ws.max_row
            for col in range(1, 3):
                self._style_data_cell(ws.cell(row=row, column=col))

        self._auto_width(ws)
